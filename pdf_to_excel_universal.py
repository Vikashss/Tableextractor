"""
=============================================================================
  Universal PDF Table Extractor
  Handles all 3 PDF types:
    1. Digital PDF       — text layer exists, pdfplumber extracts directly
    2. Scanned PDF       — image-only, OCR via pytesseract
    3. Mixed / Complex   — tries digital first, falls back to OCR per page
=============================================================================

  USAGE:
    python pdf_to_excel_universal.py <path_to_pdf> [output.xlsx]

  REQUIREMENTS:
    pip install pdfplumber openpyxl pandas pytesseract pdf2image pillow

  SYSTEM DEPENDENCIES:
    - Tesseract OCR:  sudo apt install tesseract-ocr   (Linux)
                      brew install tesseract            (Mac)
                      https://github.com/UB-Mannheim/tesseract/wiki  (Windows)
    - Poppler:        sudo apt install poppler-utils    (Linux)
                      brew install poppler              (Mac)
=============================================================================
"""

import sys
import os
import re
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Try importing OCR tools (only needed for scanned PDFs) ──────────────────
try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — DETECT PDF TYPE
# ══════════════════════════════════════════════════════════════════════════════

def detect_pdf_type(pdf_path: str) -> str:
    """
    Returns one of: 'digital', 'scanned', 'mixed'
    Logic:
      - If >50% of pages have extractable text → 'digital'
      - If <10% of pages have extractable text → 'scanned'
      - Otherwise → 'mixed'
    """
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        pages_with_text = 0
        for page in pdf.pages:
            text = page.extract_text() or ""
            if len(text.strip()) > 30:        # at least 30 chars = real text
                pages_with_text += 1

    ratio = pages_with_text / total if total > 0 else 0

    if ratio >= 0.5:
        return "digital"
    elif ratio < 0.1:
        return "scanned"
    else:
        return "mixed"


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2a — EXTRACT FROM DIGITAL PDF
# ══════════════════════════════════════════════════════════════════════════════

def extract_digital(pdf_path: str) -> list[dict]:
    """
    Uses pdfplumber to extract all tables from a digital PDF.
    Returns a list of dicts: {page, table_index, dataframe}
    """
    results = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for t_idx, raw_table in enumerate(tables, start=1):
                if not raw_table or len(raw_table) < 2:
                    continue

                # Clean None values
                cleaned = [[cell or "" for cell in row] for row in raw_table]

                # Use first row as header if it looks like one
                header = cleaned[0]
                data_rows = cleaned[1:]

                # Deduplicate column names (pdfplumber can return dupes)
                seen = {}
                unique_header = []
                for col in header:
                    col = str(col).strip() or f"col_{len(unique_header)}"
                    if col in seen:
                        seen[col] += 1
                        col = f"{col}_{seen[col]}"
                    else:
                        seen[col] = 0
                    unique_header.append(col)

                df = pd.DataFrame(data_rows, columns=unique_header)
                df = df.replace("", pd.NA).dropna(how="all").fillna("")
                df = df[df.apply(lambda r: r.str.strip().ne("").any(), axis=1)]

                results.append({
                    "page": page_num,
                    "table_index": t_idx,
                    "label": f"Page {page_num} — Table {t_idx}",
                    "dataframe": df,
                })

    return results


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2b — EXTRACT FROM SCANNED PDF (OCR)
# ══════════════════════════════════════════════════════════════════════════════

def ocr_page_to_text(image, lang: str = "eng") -> str:
    """Run Tesseract OCR on a PIL image and return extracted text."""
    config = r"--oem 3 --psm 6"
    return pytesseract.image_to_string(image, lang=lang, config=config)


def ocr_page_to_dataframe(image, lang: str = "eng") -> pd.DataFrame | None:
    """
    Use Tesseract's TSV output to reconstruct table structure from a scanned page.
    Groups text by (block_num, line_num) and attempts column alignment via x-position.
    """
    config = r"--oem 3 --psm 6"
    tsv = pytesseract.image_to_data(
        image, lang=lang, config=config,
        output_type=pytesseract.Output.DATAFRAME
    )

    # Keep only real words
    tsv = tsv[tsv["conf"] > 20].copy()
    tsv = tsv[tsv["text"].notna() & tsv["text"].str.strip().ne("")]

    if tsv.empty:
        return None

    # Group words into lines
    tsv["line_id"] = tsv["block_num"].astype(str) + "_" + tsv["par_num"].astype(str) + "_" + tsv["line_num"].astype(str)
    lines = tsv.groupby("line_id", sort=False).apply(
        lambda g: g.sort_values("left")[["left", "text"]].values.tolist()
    ).tolist()

    if not lines:
        return None

    # Detect column x-boundaries using the first few data rows
    # Simple heuristic: cluster x-positions across lines
    all_x = [word[0] for line in lines[:10] for word in line]
    if not all_x:
        return None

    col_boundaries = cluster_x_positions(all_x)

    # Assign each word to a column bucket
    rows = []
    for line in lines:
        row = [""] * len(col_boundaries)
        for x, text in line:
            col_idx = assign_column(x, col_boundaries)
            if 0 <= col_idx < len(row):
                row[col_idx] = (row[col_idx] + " " + text).strip()
        rows.append(row)

    df = pd.DataFrame(rows)
    df = df.replace("", pd.NA).dropna(how="all").fillna("")
    df = df[df.apply(lambda r: r.str.strip().ne("").any(), axis=1)]

    # Promote first row to header if it has text
    if not df.empty:
        header = df.iloc[0].tolist()
        # Only use as header if most cells look like labels (not numbers)
        label_like = sum(1 for h in header if h and not re.match(r"^\d+$", str(h)))
        if label_like >= len(header) * 0.5:
            df.columns = [str(h) or f"col_{i}" for i, h in enumerate(header)]
            df = df.iloc[1:].reset_index(drop=True)

    return df if not df.empty else None


def cluster_x_positions(x_positions: list, tolerance: int = 30) -> list:
    """Simple greedy clustering of x-positions into column boundaries."""
    if not x_positions:
        return [0]
    sorted_x = sorted(set(x_positions))
    clusters = [sorted_x[0]]
    for x in sorted_x[1:]:
        if x - clusters[-1] > tolerance:
            clusters.append(x)
    return clusters


def assign_column(x: int, boundaries: list) -> int:
    """Find which column bucket an x-position belongs to."""
    for i in range(len(boundaries) - 1):
        if boundaries[i] <= x < boundaries[i + 1]:
            return i
    return len(boundaries) - 1


def extract_scanned(pdf_path: str, dpi: int = 250, lang: str = "eng") -> list[dict]:
    """
    Converts each page to an image, runs OCR, and extracts tables.
    Returns same format as extract_digital().
    """
    if not OCR_AVAILABLE:
        raise RuntimeError(
            "OCR libraries not installed.\n"
            "Run: pip install pytesseract pdf2image pillow\n"
            "Also install Tesseract: https://tesseract-ocr.github.io/tessdoc/Installation.html"
        )

    print(f"  Converting PDF to images at {dpi} DPI...")
    images = convert_from_path(pdf_path, dpi=dpi)
    print(f"  {len(images)} page(s) detected. Running OCR...")

    results = []
    for page_num, img in enumerate(images, start=1):
        print(f"    OCR → page {page_num}/{len(images)}")

        df = ocr_page_to_dataframe(img, lang=lang)
        if df is not None and not df.empty:
            results.append({
                "page": page_num,
                "table_index": 1,
                "label": f"Page {page_num} (OCR)",
                "dataframe": df,
            })

    return results


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2c — EXTRACT FROM MIXED PDF
# ══════════════════════════════════════════════════════════════════════════════

def extract_mixed(pdf_path: str, dpi: int = 250, lang: str = "eng") -> list[dict]:
    """
    Page-by-page: tries pdfplumber first, falls back to OCR if no text found.
    """
    if not OCR_AVAILABLE:
        print("  Warning: OCR libraries not available. Only digital pages will be extracted.")

    results = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        images = None  # Lazy-load only if OCR needed

        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            has_text = len(text.strip()) > 30

            if has_text:
                # Digital page — use pdfplumber
                tables = page.extract_tables()
                for t_idx, raw_table in enumerate(tables, start=1):
                    if not raw_table or len(raw_table) < 2:
                        continue
                    cleaned = [[cell or "" for cell in row] for row in raw_table]
                    header = [str(h).strip() or f"col_{i}" for i, h in enumerate(cleaned[0])]
                    df = pd.DataFrame(cleaned[1:], columns=header)
                    df = df.replace("", pd.NA).dropna(how="all").fillna("")
                    results.append({
                        "page": page_num,
                        "table_index": t_idx,
                        "label": f"Page {page_num} — Table {t_idx} (digital)",
                        "dataframe": df,
                    })
            else:
                # Scanned page — fall back to OCR
                if OCR_AVAILABLE:
                    if images is None:
                        print(f"  Some pages need OCR. Converting full PDF to images...")
                        images = convert_from_path(pdf_path, dpi=dpi)
                    img = images[page_num - 1]
                    print(f"    OCR → page {page_num}/{total_pages}")
                    df = ocr_page_to_dataframe(img, lang=lang)
                    if df is not None and not df.empty:
                        results.append({
                            "page": page_num,
                            "table_index": 1,
                            "label": f"Page {page_num} (OCR fallback)",
                            "dataframe": df,
                        })

    return results


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3 — WRITE TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

# Styles
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
ALT_FILL  = PatternFill("solid", fgColor="D6E4F0")
TAB_FILL  = PatternFill("solid", fgColor="2E75B6")
TAB_FONT  = Font(bold=True, color="FFFFFF", size=12, name="Arial")
DATA_FONT = Font(size=10, name="Arial")
_thin     = Side(style="thin", color="B0B0B0")
BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP      = Alignment(wrap_text=True, vertical="top")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_col_width(df: pd.DataFrame) -> list[int]:
    """Estimate reasonable column widths from content."""
    widths = []
    for col in df.columns:
        series = df[col].astype(str)
        lengths = series.apply(len)
        max_data_len = int(lengths.max()) if not df.empty else 0
        max_len = max(len(str(col)), max_data_len)
        widths.append(min(max(max_len + 2, 8), 50))
    return widths


def write_table_to_sheet(wb: openpyxl.Workbook, table: dict):
    """Write a single extracted table to its own Excel sheet."""
    label = table["label"][:31]   # Excel sheet name max = 31 chars
    df    = table["dataframe"]

    # Avoid duplicate sheet names
    base = label
    counter = 1
    while label in wb.sheetnames:
        label = f"{base[:27]}_{counter}"
        counter += 1

    ws = wb.create_sheet(title=label)
    ws.sheet_view.showGridLines = False

    num_cols = len(df.columns)

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    title_cell = ws.cell(row=1, column=1, value=table["label"])
    title_cell.fill      = TAB_FILL
    title_cell.font      = TAB_FONT
    title_cell.alignment = CENTER
    title_cell.border    = BORDER
    ws.row_dimensions[1].height = 22

    # ── Header row ───────────────────────────────────────────────────────────
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=str(col))
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = CENTER
        c.border    = BORDER
    ws.row_dimensions[2].height = 28

    # ── Data rows ────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        alt = (ri % 2 == 0)
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            c.font      = DATA_FONT
            c.border    = BORDER
            c.alignment = WRAP
            if alt:
                c.fill = ALT_FILL
        ws.row_dimensions[ri].height = 18

    # ── Column widths ────────────────────────────────────────────────────────
    for ci, w in enumerate(auto_col_width(df), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_summary_sheet(wb: openpyxl.Workbook, tables: list[dict], pdf_path: str, pdf_type: str):
    """Write a summary index as the first sheet."""
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="PDF Table Extraction — Summary")
    t.fill      = TAB_FILL
    t.font      = TAB_FONT
    t.alignment = CENTER
    t.border    = BORDER
    ws.row_dimensions[1].height = 26

    # Info rows
    info = [
        ("Source file",  os.path.basename(pdf_path)),
        ("PDF type",     pdf_type.upper()),
        ("Tables found", len(tables)),
        ("Total rows",   sum(len(t["dataframe"]) for t in tables)),
    ]
    for ri, (k, v) in enumerate(info, start=2):
        label = ws.cell(row=ri, column=1, value=k)
        value = ws.cell(row=ri, column=2, value=str(v))
        label.font = Font(bold=True, size=10, name="Arial")
        value.font = DATA_FONT
        label.border = value.border = BORDER
        ws.row_dimensions[ri].height = 16

    # Table index
    ws.cell(row=7, column=1, value="").border = BORDER
    headers = ["#", "Sheet name", "Page", "Table", "Rows", "Columns"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[8].height = 22

    for ri, tbl in enumerate(tables, start=9):
        alt = (ri % 2 == 0)
        vals = [
            ri - 8,
            tbl["label"][:31],
            tbl["page"],
            tbl["table_index"],
            len(tbl["dataframe"]),
            len(tbl["dataframe"].columns),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = DATA_FONT
            c.border = BORDER
            c.alignment = CENTER
            if alt:
                c.fill = ALT_FILL
        ws.row_dimensions[ri].height = 16

    col_widths = [5, 35, 8, 8, 8, 8]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def save_to_excel(tables: list[dict], output_path: str, pdf_path: str, pdf_type: str):
    """Write all extracted tables to an Excel workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # Remove default empty sheet

    write_summary_sheet(wb, tables, pdf_path, pdf_type)

    for table in tables:
        write_table_to_sheet(wb, table)

    wb.save(output_path)
    print(f"\n✓ Saved: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)}  (1 summary + {len(tables)} table sheets)")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def extract_pdf_to_excel(
    pdf_path: str,
    output_path: str = None,
    force_type: str = None,
    dpi: int = 250,
    lang: str = "eng",
) -> str:
    """
    Main function. Auto-detects PDF type and extracts all tables to Excel.

    Parameters
    ----------
    pdf_path    : path to the input PDF
    output_path : path for the output .xlsx (defaults to <pdf_name>_tables.xlsx)
    force_type  : override auto-detection — 'digital', 'scanned', or 'mixed'
    dpi         : DPI for image conversion (scanned/mixed only); higher = better quality
    lang        : Tesseract language code(s), e.g. 'eng', 'hin', 'eng+hin'

    Returns
    -------
    Path to the saved Excel file.
    """
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    if output_path is None:
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        output_path = os.path.join(os.path.dirname(pdf_path), f"{base}_tables.xlsx")

    print(f"\n{'='*60}")
    print(f"  PDF Table Extractor — Universal")
    print(f"{'='*60}")
    print(f"  Input : {pdf_path}")
    print(f"  Output: {output_path}")

    # ── Detect type ──────────────────────────────────────────────────────────
    if force_type:
        pdf_type = force_type.lower()
        print(f"  Type  : {pdf_type.upper()} (forced)")
    else:
        pdf_type = detect_pdf_type(pdf_path)
        print(f"  Type  : {pdf_type.upper()} (auto-detected)")

    # ── Extract ──────────────────────────────────────────────────────────────
    print(f"\nExtracting tables...")

    if pdf_type == "digital":
        tables = extract_digital(pdf_path)

    elif pdf_type == "scanned":
        tables = extract_scanned(pdf_path, dpi=dpi, lang=lang)

    elif pdf_type == "mixed":
        tables = extract_mixed(pdf_path, dpi=dpi, lang=lang)

    else:
        raise ValueError(f"Unknown PDF type '{pdf_type}'. Use: digital / scanned / mixed")

    # ── Report ───────────────────────────────────────────────────────────────
    if not tables:
        print("\n⚠ No tables found in this PDF.")
        print("  Tips:")
        print("  - For scanned PDFs, try a higher DPI: dpi=300 or dpi=400")
        print("  - For non-English PDFs, specify lang= (e.g., lang='hin+eng')")
        print("  - Try force_type='scanned' if auto-detection was wrong")
        return None

    total_rows = sum(len(t["dataframe"]) for t in tables)
    print(f"\n  Found {len(tables)} table(s) with {total_rows} total data rows")

    # ── Write Excel ──────────────────────────────────────────────────────────
    save_to_excel(tables, output_path, pdf_path, pdf_type)
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        print("USAGE:  python pdf_to_excel_universal.py input.pdf [output.xlsx]")
        print()
        print("OPTIONAL — override auto-detection:")
        print("  Add  --type=digital | scanned | mixed")
        print("  Add  --dpi=300   (default 250, increase for blurry scans)")
        print("  Add  --lang=hin  (Tesseract language; default 'eng')")
        sys.exit(1)

    pdf_file    = sys.argv[1]
    output_file = None
    force_type  = None
    dpi         = 250
    lang        = "eng"

    for arg in sys.argv[2:]:
        if arg.startswith("--type="):
            force_type = arg.split("=", 1)[1]
        elif arg.startswith("--dpi="):
            dpi = int(arg.split("=", 1)[1])
        elif arg.startswith("--lang="):
            lang = arg.split("=", 1)[1]
        elif not arg.startswith("--"):
            output_file = arg

    extract_pdf_to_excel(
        pdf_path=pdf_file,
        output_path=output_file,
        force_type=force_type,
        dpi=dpi,
        lang=lang,
    )
